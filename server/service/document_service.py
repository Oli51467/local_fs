import hashlib
import re
from pathlib import Path
from typing import List, Dict, Any, Tuple
import numpy as np

# 文档处理库
import pdfplumber
from docx import Document
import openpyxl
from sentence_transformers import SentenceTransformer
import jieba

from config.config import DocumentConfig

class DocumentProcessor:
    """文档处理器"""
    
    def __init__(self):
        # 初始化向量化模型
        self.embedding_model = SentenceTransformer(DocumentConfig.EMBEDDING_MODEL)
        
    def extract_text(self, file_path: str, file_type: str) -> str:
        """从文件中提取文本内容"""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        try:
            if file_type.lower() == '.txt':
                return self._extract_txt(file_path)
            elif file_type.lower() == '.md':
                return self._extract_txt(file_path)  # Markdown按文本处理
            elif file_type.lower() == '.pdf':
                return self._extract_pdf(file_path)
            elif file_type.lower() in ['.docx', '.doc']:
                return self._extract_docx(file_path)
            elif file_type.lower() in ['.xlsx', '.xls']:
                return self._extract_xlsx(file_path)
            else:
                raise ValueError(f"不支持的文件类型: {file_type}")
        except Exception as e:
            raise Exception(f"文本提取失败: {str(e)}")
    
    def _extract_txt(self, file_path: Path) -> str:
        """提取TXT文件内容"""
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
        raise Exception("无法解码文件，尝试了多种编码格式")
    
    def _extract_pdf(self, file_path: Path) -> str:
        """提取PDF文件内容"""
        text = ""
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text.strip()
    
    def _extract_docx(self, file_path: Path) -> str:
        """提取DOCX文件内容"""
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    
    def _extract_xlsx(self, file_path: Path) -> str:
        """提取Excel文件内容"""
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"工作表: {sheet_name}\n"
            
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
            text += "\n"
        
        return text.strip()
    
    def split_text_into_chunks(self, text: str, chunk_size: int = None, 
                              chunk_overlap: int = None) -> List[str]:
        """将文本分割成块"""
        if chunk_size is None:
            chunk_size = DocumentConfig.CHUNK_SIZE
        if chunk_overlap is None:
            chunk_overlap = DocumentConfig.CHUNK_OVERLAP
        
        # 清理文本
        text = self._clean_text(text)
        
        if len(text) <= chunk_size:
            return [text] if text.strip() else []
        
        chunks = []
        start = 0
        
        while start < len(text):
            end = start + chunk_size
            
            # 如果不是最后一块，尝试在句号、换行符等处分割
            if end < len(text):
                # 寻找合适的分割点
                for delimiter in ['。', '\n', '！', '？', '.', '!', '?']:
                    last_delimiter = text.rfind(delimiter, start, end)
                    if last_delimiter != -1:
                        end = last_delimiter + 1
                        break
            
            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            
            # 计算下一个块的起始位置（考虑重叠）
            start = max(start + 1, end - chunk_overlap)
            
            # 避免无限循环
            if start >= len(text):
                break
        
        return chunks
    
    def _clean_text(self, text: str) -> str:
        """清理文本"""
        # 移除多余的空白字符
        text = re.sub(r'\s+', ' ', text)
        # 移除特殊字符（保留中文、英文、数字、基本标点）
        text = re.sub(r'[^\u4e00-\u9fff\w\s.,!?;:()\[\]{}"\'-]', '', text)
        return text.strip()
    
    def vectorize_text(self, texts: List[str]) -> np.ndarray:
        """将文本向量化"""
        if not texts:
            return np.array([])
        
        try:
            # 清理文本，移除可能导致问题的字符
            cleaned_texts = []
            for text in texts:
                if text and text.strip():
                    # 移除特殊字符，限制长度
                    cleaned_text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\xff]', '', text)
                    cleaned_text = cleaned_text[:1000]  # 限制最大长度
                    if cleaned_text.strip():
                        cleaned_texts.append(cleaned_text.strip())
            
            if not cleaned_texts:
                return np.array([])
            
            # 分批处理，避免内存问题
            batch_size = 32
            all_embeddings = []
            
            for i in range(0, len(cleaned_texts), batch_size):
                batch_texts = cleaned_texts[i:i + batch_size]
                try:
                    # 使用sentence-transformers进行向量化
                    batch_embeddings = self.embedding_model.encode(
                        batch_texts, 
                        convert_to_numpy=True,
                        show_progress_bar=False,
                        batch_size=min(16, len(batch_texts))
                    )
                    all_embeddings.append(batch_embeddings)
                except Exception as e:
                    print(f"批处理向量化失败: {e}")
                    # 如果批处理失败，尝试逐个处理
                    for text in batch_texts:
                        try:
                            single_embedding = self.embedding_model.encode(
                                [text], 
                                convert_to_numpy=True,
                                show_progress_bar=False
                            )
                            all_embeddings.append(single_embedding)
                        except Exception as single_e:
                            print(f"单个文本向量化失败: {single_e}")
                            # 创建零向量作为fallback
                            zero_embedding = np.zeros((1, DocumentConfig.VECTOR_DIMENSION))
                            all_embeddings.append(zero_embedding)
            
            if all_embeddings:
                return np.vstack(all_embeddings)
            else:
                return np.array([])
                
        except Exception as e:
            print(f"向量化过程出错: {e}")
            # 返回零向量作为fallback
            return np.zeros((len(texts), DocumentConfig.VECTOR_DIMENSION))
    
    def calculate_file_hash(self, file_path: str) -> str:
        """计算文件哈希值"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    
    def process_document(self, file_path: str, filename: str = None) -> Dict[str, Any]:
        """处理文档：提取文本、分段、向量化"""
        file_path = Path(file_path)
        
        if filename is None:
            filename = file_path.name
        
        # 获取文件信息
        file_type = file_path.suffix.lower()
        file_size = file_path.stat().st_size
        content_hash = self.calculate_file_hash(str(file_path))
        
        # 检查文件类型是否支持
        if file_type not in DocumentConfig.SUPPORTED_EXTENSIONS:
            raise ValueError(f"不支持的文件类型: {file_type}")
        
        # 提取文本
        text_content = self.extract_text(str(file_path), file_type)
        
        if not text_content.strip():
            raise ValueError("文档内容为空")
        
        # 分割文本
        chunks = self.split_text_into_chunks(text_content)
        
        if not chunks:
            raise ValueError("文档分割后无有效内容")
        
        # 向量化
        vectors = self.vectorize_text(chunks)
        
        return {
            'filename': filename,
            'file_path': str(file_path),
            'file_type': file_type,
            'file_size': file_size,
            'content_hash': content_hash,
            'text_content': text_content,
            'chunks': chunks,
            'vectors': vectors,
            'total_chunks': len(chunks)
        }
    
    def search_text_similarity(self, query: str, texts: List[str], top_k: int = 5) -> List[Tuple[int, float]]:
        """计算查询文本与文本列表的相似度"""
        if not texts:
            return []
        
        # 向量化查询和文本
        query_vector = self.embedding_model.encode([query], convert_to_numpy=True)[0]
        text_vectors = self.embedding_model.encode(texts, convert_to_numpy=True)
        
        # 计算余弦相似度
        similarities = np.dot(text_vectors, query_vector) / (
            np.linalg.norm(text_vectors, axis=1) * np.linalg.norm(query_vector)
        )
        
        # 获取top_k结果
        top_indices = np.argsort(similarities)[::-1][:top_k]
        results = [(int(idx), float(similarities[idx])) for idx in top_indices]
        
        return results